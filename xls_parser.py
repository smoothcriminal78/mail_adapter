from io import BytesIO
import pandas as pd
import sys
import os
import psycopg2
import datetime
from openpyxl import load_workbook
from config import *

def save_to_file(goods):
    f=open(os.path.join(os.getcwd(), 'goods'), 'a', encoding='UTF-8')
    print('\n'.join(['{} {} {}'.format(k,*v) for k,v in goods.items()]), file=f)

def save_to_db(goods):
    conn = psycopg2.connect(database = DB_NAME, user = DB_USER, password = DB_PASSWORD, host = DB_HOST, port = DB_PORT)
    print("Opened database successfully")
    cur = conn.cursor()

    for k,v in goods.items():
        product, qty, price = k, *v
        cur.execute('''INSERT INTO catalogapplic_product (product_name, quantity, price, pub_date) VALUES (%s, %s, %s, %s);''', (product, int(qty), int(price), datetime.datetime.now()))

    conn.commit()
    conn.close()

def parse(data):
    wb = load_workbook(filename=BytesIO(data))
    # wb = load_workbook('attachment.xlsx')
    sheet = wb.get_sheet_by_name('Лист1')
    df = pd.DataFrame(sheet.values)

    # product -> (qty, price)
    goods = {row[0] : (row[1], row[2]) for i, row in df[1:].iterrows()}
    # save_to_file(goods)
    save_to_db(goods)
